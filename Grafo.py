import openpyxl
from geopy.distance import geodesic
import matplotlib.pyplot as plt
import networkx as nx
import tkinter as tk

# Caminho para o arquivo Excel no repositório clonado
excel_file_path = "Pasta10.xlsx"

# Carregar o arquivo Excel usando openpyxl
workbook = openpyxl.load_workbook(excel_file_path)
sheet = workbook.active


class Grafo:
    def __init__(self):
        self.grafo = {}

    # Adiciona uma aresta ao grafo
    def adiciona_aresta(self, u, v, w):
        if u not in self.grafo:
            self.grafo[u] = []
        self.grafo[u].append((v, w))
        if v not in self.grafo:
            self.grafo[v] = []

    def imprime_caminho(self, pred, src, dest):
        caminho = []
        vertice = dest
        while vertice != src:
            caminho.append(vertice)
            vertice = pred[vertice]
        caminho.append(src)
        caminho.reverse()
        output_text.set("Caminho: " + " -> ".join(caminho))

        output_text.set(output_text.get() + "\n\nPeso das arestas:\n")
        total_weight = 0
        for i in range(len(caminho) - 1):
            u = caminho[i]
            v = caminho[i + 1]
            for neighbor, weight in self.grafo[u]:
                if neighbor == v:
                    output_text.set(output_text.get() + f"\n{u} -> {v}: {weight:.2f} km")
                    total_weight += weight
                    break

        output_text.set(output_text.get() + f"\n\nPeso total: {total_weight:.2f} km")

        # Visualizar o grafo
        self.visualizar_grafo(list(self.grafo.keys()), arestas, src, dest)

    # O algoritmo de Bellman-Ford
    def bellman_ford(self, src, dest):
        # Inicializa as distâncias do vértice inicial como infinito
        dist = {vertice: float("Inf") for vertice in self.grafo}
        pred = {vertice: None for vertice in self.grafo}
        dist[src] = 0

        # Relaxa todas as arestas |V| - 1 vezes
        for _ in range(len(self.grafo) - 1):
            for u in self.grafo:
                for v, w in self.grafo[u]:
                    if dist[u] != float("Inf") and dist[u] + w < dist[v]:
                        dist[v] = dist[u] + w
                        pred[v] = u

        output_text.set(f"O menor caminho entre {src} e {dest} é:")
        self.imprime_caminho(pred, src, dest)

        return pred, dist

    # Método para visualizar o grafo
    def visualizar_grafo(self, vertices, arestas, src, dest):
        G = nx.DiGraph()

        for vertice in vertices:
            G.add_node(vertice)

        for u, v, w in arestas:
            G.add_edge(u, v, weight=w)

        pos = nx.spring_layout(G)
        nx.draw(G, pos, with_labels=True, node_size=650, node_color="yellow", font_size=7, font_weight="bold",
                width=1, edge_color="gray", arrows=True)

        # Destaca o caminho entre src e dest em vermelho
        if src and dest:
            shortest_path_nodes = nx.bellman_ford_path(G, source=src, target=dest, weight='weight')
            highlighted_edges = [(shortest_path_nodes[i], shortest_path_nodes[i + 1]) for i in
                                 range(len(shortest_path_nodes) - 1)]
            nx.draw_networkx_edges(G, pos, edgelist=highlighted_edges, edge_color='red', width=2.0)

        plt.title("Grafo")
        plt.show()


def calcular_caminho():
    src = entry_src.get()
    dest = entry_dest.get()
    pred, dist = grafo.bellman_ford(src, dest)


# Função para lidar com o fechamento da janela
def fechar_janela():
    root.destroy()


# Interface gráfica
root = tk.Tk()
root.title("Calculadora de Menor Caminho")

frame = tk.Frame(root)
frame.pack(padx=20, pady=20)

label_src = tk.Label(frame, text="Origem:")
label_src.grid(row=0, column=0, padx=5, pady=5, sticky="e")

entry_src = tk.Entry(frame)
entry_src.grid(row=0, column=1, padx=5, pady=5)

label_dest = tk.Label(frame, text="Destino:")
label_dest.grid(row=1, column=0, padx=5, pady=5, sticky="e")

entry_dest = tk.Entry(frame)
entry_dest.grid(row=1, column=1, padx=5, pady=5)

button_calcular = tk.Button(frame, text="Calcular Caminho", command=calcular_caminho)
button_calcular.grid(row=2, columnspan=2, padx=5, pady=5)

output_text = tk.StringVar()
output_label = tk.Label(frame, textvariable=output_text, justify="left", font=("Helvetica", 12))
output_label.grid(row=3, columnspan=2, padx=5, pady=5, sticky="w")

# Tratamento de evento para o fechamento da janela
root.protocol("WM_DELETE_WINDOW", fechar_janela)

grafo = Grafo()
arestas = []

for d in range(94):
    celula = sheet.cell(row=(d + 2), column=1)
    bairro1 = celula.value
    celula = sheet.cell(row=(d + 2), column=3)
    coordenadas1 = celula.value.split(", ")
    X1 = float(coordenadas1[0])
    Y1 = float(coordenadas1[1])
    coord_inicial = (X1, Y1)
    celula = sheet.cell(row=(d + 2), column=2)
    limitrofes = celula.value.split(", ")
    for limitrofe in limitrofes:
        linha = None
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
            for cell in row:
                if cell.value == limitrofe:
                    linha = cell.row
                    break
        celula = sheet.cell(row=linha, column=3)
        coordenadas2 = celula.value.split(", ")
        X2 = float(coordenadas2[0])
        Y2 = float(coordenadas2[1])
        coord_final = (X2, Y2)
        peso = geodesic(coord_inicial, coord_final).kilometers
        arestas.append((bairro1, limitrofe, peso))
        grafo.adiciona_aresta(bairro1, limitrofe, peso)

root.mainloop()